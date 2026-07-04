# SPDX-FileCopyrightText: 2026 Weibo, Inc.
#
# SPDX-License-Identifier: Apache-2.0

"""Row-aware Q&A unitization for FAQ-like Excel workbooks."""

from __future__ import annotations

import re
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Any, Iterable

from llama_index.core.schema import TextNode
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

MIN_FAQ_ROWS = 2
MIN_QA_COVERAGE = 0.35
MIN_COLUMN_CONFIDENCE = 0.60
MIN_SHEET_CONFIDENCE = 0.65
HEADER_SCAN_ROWS = 10
VALUE_SAMPLE_ROWS = 200
EXCEL_FAQ_PARSER_SUBTYPE = "qa_pair"
EXCEL_FAQ_SOURCE_FORMAT = "excel_faq"

QUESTION_NAMES = {
    "q",
    "question",
    "问题",
    "标准q",
    "标准问",
    "主问题",
    "用户问题",
    "咨询问题",
    "问法",
}
ALTERNATE_QUESTION_NAMES = {
    "相似q",
    "相似问",
    "扩展问",
    "关联问",
    "近似问",
    "同义问",
    "alternatequestions",
    "aliases",
}
ANSWER_NAMES = {
    "a",
    "answer",
    "答案",
    "内容",
    "回复",
    "答复",
    "话术",
    "处理方式",
    "解决方案",
}
CONTEXT_NAME_HINTS = {
    "分类",
    "场景",
    "业务线",
    "模块",
    "标签",
    "渠道",
    "适用对象",
    "流程",
    "来源",
    "状态",
    "时间",
}
ALTERNATE_SPLIT_RE = re.compile(r"[;；|/\n]+")


@dataclass(frozen=True, slots=True)
class ExcelFAQColumnMapping:
    question_column: str
    answer_columns: list[str]
    alternate_question_columns: list[str]
    context_columns: list[str]
    confidence: float
    question_confidence: float
    answer_confidence: float


@dataclass(frozen=True, slots=True)
class ExcelFAQSheetSummary:
    sheet_name: str
    header_row: int
    data_row_count: int
    qa_row_count: int
    skipped_row_count: int
    confidence: float
    mapping: ExcelFAQColumnMapping


@dataclass(frozen=True, slots=True)
class ExcelFAQUnit:
    sheet_name: str
    row_number: int
    question: str
    alternate_questions: list[str]
    answer: str
    context_fields: dict[str, str]
    source_position: str
    confidence: float


@dataclass(frozen=True, slots=True)
class ExcelFAQUnitizationResult:
    detected: bool
    confidence: float
    nodes: list[TextNode]
    sheet_summaries: list[ExcelFAQSheetSummary]
    qa_row_count: int
    skipped_row_count: int


@dataclass(frozen=True, slots=True)
class _SheetTable:
    header_row: int
    headers: list[str]
    rows: list[tuple[int, dict[str, str]]]


def build_excel_qa_pair_nodes_from_binary(
    binary_data: bytes,
    *,
    source_file: str,
) -> ExcelFAQUnitizationResult | None:
    workbook = load_workbook(
        BytesIO(binary_data),
        data_only=True,
        read_only=True,
    )
    return _build_excel_qa_pair_nodes(workbook, source_file=source_file)


def build_excel_qa_pair_nodes_from_file(
    file_path: str,
    *,
    source_file: str | None = None,
) -> ExcelFAQUnitizationResult | None:
    workbook = load_workbook(
        file_path,
        data_only=True,
        read_only=True,
    )
    return _build_excel_qa_pair_nodes(
        workbook,
        source_file=source_file or Path(file_path).name,
    )


def _build_excel_qa_pair_nodes(
    workbook,
    *,
    source_file: str,
) -> ExcelFAQUnitizationResult | None:
    all_units: list[ExcelFAQUnit] = []
    sheet_summaries: list[ExcelFAQSheetSummary] = []
    skipped_rows = 0

    for worksheet in workbook.worksheets:
        if worksheet.sheet_state != "visible":
            continue

        table = _extract_sheet_table(worksheet)
        if table is None:
            continue

        mapping = _detect_column_mapping(table)
        if mapping is None:
            continue

        units, skipped = _build_units_for_sheet(
            worksheet=worksheet,
            table=table,
            mapping=mapping,
        )
        skipped_rows += skipped
        if not units:
            continue

        data_row_count = len(table.rows)
        qa_row_count = len(units)
        coverage = qa_row_count / data_row_count if data_row_count else 0
        sheet_confidence = min(
            1.0,
            (mapping.confidence * 0.75) + (min(coverage, 1.0) * 0.25),
        )
        if (
            qa_row_count < MIN_FAQ_ROWS
            or coverage < MIN_QA_COVERAGE
            or sheet_confidence < MIN_SHEET_CONFIDENCE
        ):
            skipped_rows += qa_row_count
            continue

        all_units.extend(units)
        sheet_summaries.append(
            ExcelFAQSheetSummary(
                sheet_name=worksheet.title,
                header_row=table.header_row,
                data_row_count=data_row_count,
                qa_row_count=qa_row_count,
                skipped_row_count=skipped,
                confidence=sheet_confidence,
                mapping=mapping,
            )
        )

    if not all_units or not sheet_summaries:
        return None

    workbook_confidence = sum(summary.confidence for summary in sheet_summaries) / len(
        sheet_summaries
    )
    nodes = [
        _build_text_node(
            unit,
            qa_index=index,
            source_file=source_file,
            workbook_confidence=workbook_confidence,
        )
        for index, unit in enumerate(all_units)
    ]
    return ExcelFAQUnitizationResult(
        detected=True,
        confidence=workbook_confidence,
        nodes=nodes,
        sheet_summaries=sheet_summaries,
        qa_row_count=len(all_units),
        skipped_row_count=skipped_rows,
    )


def _extract_sheet_table(worksheet: Worksheet) -> _SheetTable | None:
    header_row, headers = _detect_header_row(worksheet)
    if header_row is None:
        return None

    rows: list[tuple[int, dict[str, str]]] = []
    for row_index, row in enumerate(
        worksheet.iter_rows(min_row=header_row + 1, values_only=True),
        start=header_row + 1,
    ):
        values = {
            header: _cell_to_text(value)
            for header, value in zip(headers, row)
            if header and _cell_to_text(value)
        }
        if values:
            rows.append((row_index, values))

    if not rows:
        return None

    return _SheetTable(
        header_row=header_row,
        headers=headers,
        rows=rows,
    )


def _detect_header_row(worksheet: Worksheet) -> tuple[int | None, list[str]]:
    best_row: int | None = None
    best_headers: list[str] = []
    best_score = 0.0

    max_row = min(worksheet.max_row or 0, HEADER_SCAN_ROWS)
    for row_index, row in enumerate(
        worksheet.iter_rows(min_row=1, max_row=max_row, values_only=True),
        start=1,
    ):
        headers = [_cell_to_text(value) for value in row]
        non_empty_headers = [header for header in headers if header]
        if len(non_empty_headers) < 2:
            continue

        normalized_headers = [_normalize_header(header) for header in non_empty_headers]
        name_hits = sum(
            1
            for header in normalized_headers
            if _name_score(header, QUESTION_NAMES) > 0
            or _name_score(header, ANSWER_NAMES) > 0
            or _name_score(header, ALTERNATE_QUESTION_NAMES) > 0
        )
        score = len(non_empty_headers) * 0.05 + name_hits
        if score > best_score:
            best_score = score
            best_row = row_index
            best_headers = _dedupe_headers(headers)

    if best_row is None or best_score < 1.0:
        return None, []

    return best_row, best_headers


def _detect_column_mapping(table: _SheetTable) -> ExcelFAQColumnMapping | None:
    sampled_rows = table.rows[:VALUE_SAMPLE_ROWS]
    scores = {
        header: _score_column(header, [row.get(header, "") for _, row in sampled_rows])
        for header in table.headers
        if header
    }
    if not scores:
        return None

    question_column, question_score = max(
        ((header, role_scores["question"]) for header, role_scores in scores.items()),
        key=lambda item: item[1],
    )
    answer_column, answer_score = max(
        ((header, role_scores["answer"]) for header, role_scores in scores.items()),
        key=lambda item: item[1],
    )
    if question_column == answer_column:
        return None
    if question_score < MIN_COLUMN_CONFIDENCE or answer_score < MIN_COLUMN_CONFIDENCE:
        return None

    alternate_question_columns = [
        header
        for header, role_scores in scores.items()
        if header not in {question_column, answer_column}
        and role_scores["alternate_question"] >= 0.55
    ]
    answer_columns = [answer_column]
    for header, role_scores in scores.items():
        if header in {question_column, answer_column, *alternate_question_columns}:
            continue
        if role_scores["answer"] >= 0.68:
            answer_columns.append(header)

    reserved_columns = {question_column, *answer_columns, *alternate_question_columns}
    context_columns = [
        header for header in table.headers if header not in reserved_columns
    ]
    confidence = min(1.0, (question_score + answer_score) / 2)
    return ExcelFAQColumnMapping(
        question_column=question_column,
        answer_columns=answer_columns,
        alternate_question_columns=alternate_question_columns,
        context_columns=context_columns,
        confidence=confidence,
        question_confidence=question_score,
        answer_confidence=answer_score,
    )


def _score_column(header: str, values: list[str]) -> dict[str, float]:
    normalized_header = _normalize_header(header)
    non_empty_values = [value for value in values if value]
    sample_count = max(len(values), 1)
    non_empty_ratio = len(non_empty_values) / sample_count
    avg_len = (
        sum(len(value) for value in non_empty_values) / len(non_empty_values)
        if non_empty_values
        else 0
    )
    unique_ratio = (
        len(set(non_empty_values)) / len(non_empty_values) if non_empty_values else 0
    )
    delimiter_ratio = (
        sum(1 for value in non_empty_values if ALTERNATE_SPLIT_RE.search(value))
        / len(non_empty_values)
        if non_empty_values
        else 0
    )

    question = (
        _name_score(normalized_header, QUESTION_NAMES) * 0.70
        + non_empty_ratio * 0.15
        + min(unique_ratio, 1.0) * 0.10
        + _short_text_score(avg_len) * 0.05
    )
    alternate_question = (
        _name_score(normalized_header, ALTERNATE_QUESTION_NAMES) * 0.75
        + delimiter_ratio * 0.15
        + non_empty_ratio * 0.10
    )
    answer = (
        _name_score(normalized_header, ANSWER_NAMES) * 0.65
        + non_empty_ratio * 0.15
        + _long_text_score(avg_len) * 0.15
        + _answer_signal_score(non_empty_values) * 0.05
    )
    context = (
        _context_name_score(normalized_header) * 0.60
        + non_empty_ratio * 0.20
        + (1.0 - min(unique_ratio, 1.0)) * 0.20
    )
    return {
        "question": min(question, 1.0),
        "alternate_question": min(alternate_question, 1.0),
        "answer": min(answer, 1.0),
        "context": min(context, 1.0),
    }


def _build_units_for_sheet(
    *,
    worksheet: Worksheet,
    table: _SheetTable,
    mapping: ExcelFAQColumnMapping,
) -> tuple[list[ExcelFAQUnit], int]:
    units: list[ExcelFAQUnit] = []
    skipped = 0
    for row_number, row in table.rows:
        question = row.get(mapping.question_column, "")
        alternate_questions = _split_alternate_questions(
            [
                row.get(column, "")
                for column in mapping.alternate_question_columns
                if row.get(column, "")
            ]
        )
        if not question and alternate_questions:
            question = alternate_questions.pop(0)

        answer = _build_answer(row, mapping.answer_columns)
        if not question or not answer:
            skipped += 1
            continue

        context_fields = {
            column: value
            for column in mapping.context_columns
            if (value := row.get(column, ""))
        }
        alternate_questions = [
            alt
            for alt in alternate_questions
            if _normalize_text(alt) != _normalize_text(question)
        ]
        units.append(
            ExcelFAQUnit(
                sheet_name=worksheet.title,
                row_number=row_number,
                question=question,
                alternate_questions=alternate_questions,
                answer=answer,
                context_fields=context_fields,
                source_position=_source_position(
                    worksheet.title,
                    row_number,
                    len(table.headers),
                ),
                confidence=mapping.confidence,
            )
        )
    return units, skipped


def _build_text_node(
    unit: ExcelFAQUnit,
    *,
    qa_index: int,
    source_file: str,
    workbook_confidence: float,
) -> TextNode:
    display_text = _build_display_text(unit)
    retrieval_text = _build_retrieval_text(unit)
    qa_id = f"{_slug(unit.sheet_name)}-r{unit.row_number}"
    metadata: dict[str, Any] = {
        "node_role": "qa_pair",
        "source_format": EXCEL_FAQ_SOURCE_FORMAT,
        "qa_id": qa_id,
        "qa_index": qa_index,
        "question": unit.question,
        "alternate_questions": unit.alternate_questions,
        "context_fields": unit.context_fields,
        "sheet_name": unit.sheet_name,
        "row_number": unit.row_number,
        "source_position": unit.source_position,
        "source_file": source_file,
        "retrieval_text": retrieval_text,
        "display_text": display_text,
        "qa_confidence": unit.confidence,
        "excel_faq_confidence": workbook_confidence,
    }
    return TextNode(
        text=display_text,
        metadata=metadata,
        excluded_embed_metadata_keys=[
            "qa_id",
            "qa_index",
            "qa_confidence",
            "excel_faq_confidence",
            "source_position",
            "retrieval_text",
            "display_text",
            "alternate_questions",
            "context_fields",
        ],
        excluded_llm_metadata_keys=[
            "qa_id",
            "qa_index",
            "qa_confidence",
            "excel_faq_confidence",
            "source_position",
            "retrieval_text",
            "display_text",
            "alternate_questions",
            "context_fields",
        ],
    )


def _build_display_text(unit: ExcelFAQUnit) -> str:
    parts = [
        f"Q: {unit.question}",
        "",
        f"A: {unit.answer}",
    ]
    context_preview = _format_context(unit.context_fields)
    if context_preview:
        parts.extend(["", f"上下文: {context_preview}"])
    parts.append(f"来源: {unit.sheet_name} 第 {unit.row_number} 行")
    return "\n".join(parts).strip()


def _build_retrieval_text(unit: ExcelFAQUnit) -> str:
    parts = [f"Question: {unit.question}"]
    if unit.alternate_questions:
        parts.append(f"Alternate questions: {'; '.join(unit.alternate_questions)}")
    context_preview = _format_context(unit.context_fields)
    if context_preview:
        parts.append(f"Context: {context_preview}")
    digest = _answer_digest(unit.answer)
    if digest:
        parts.append(f"Answer summary: {digest}")
    return "\n".join(parts)


def _build_answer(row: dict[str, str], answer_columns: list[str]) -> str:
    values = [
        (column, row.get(column, ""))
        for column in answer_columns
        if row.get(column, "")
    ]
    if not values:
        return ""
    if len(values) == 1:
        return values[0][1]
    return "\n".join(f"{column}: {value}" for column, value in values)


def _split_alternate_questions(values: Iterable[str]) -> list[str]:
    seen = set()
    result: list[str] = []
    for value in values:
        for item in ALTERNATE_SPLIT_RE.split(value):
            normalized = _normalize_text(item)
            if not normalized or normalized in seen:
                continue
            seen.add(normalized)
            result.append(item.strip())
    return result


def _detect_column_letters(column_count: int) -> tuple[str, str]:
    from openpyxl.utils import get_column_letter

    return "A", get_column_letter(max(column_count, 1))


def _source_position(sheet_name: str, row_number: int, column_count: int) -> str:
    start_col, end_col = _detect_column_letters(column_count)
    return f"{sheet_name}!{start_col}{row_number}:{end_col}{row_number}"


def _dedupe_headers(headers: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    result: list[str] = []
    for index, header in enumerate(headers, start=1):
        clean = header or f"Column{index}"
        count = seen.get(clean, 0)
        seen[clean] = count + 1
        result.append(clean if count == 0 else f"{clean}_{count + 1}")
    return result


def _name_score(normalized_header: str, names: set[str]) -> float:
    if normalized_header in names:
        return 1.0
    if any(name and name in normalized_header for name in names if len(name) >= 2):
        return 0.85
    return 0.0


def _context_name_score(normalized_header: str) -> float:
    return 1.0 if any(hint in normalized_header for hint in CONTEXT_NAME_HINTS) else 0.0


def _short_text_score(avg_len: float) -> float:
    if avg_len <= 0:
        return 0.0
    if avg_len <= 80:
        return 1.0
    if avg_len <= 160:
        return 0.5
    return 0.0


def _long_text_score(avg_len: float) -> float:
    if avg_len >= 50:
        return 1.0
    if avg_len >= 20:
        return 0.6
    return 0.0


def _answer_signal_score(values: list[str]) -> float:
    if not values:
        return 0.0
    signals = ("答案", "结论", "步骤", "流程", "http", "\n", "：", ":")
    matched = sum(1 for value in values if any(signal in value for signal in signals))
    return matched / len(values)


def _cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return re.sub(r"[ \t]+", " ", text)


def _normalize_header(value: str) -> str:
    return re.sub(r"[\s_:\-：/（）()]+", "", value.strip().lower())


def _normalize_text(value: str) -> str:
    return re.sub(r"\s+", " ", value.strip()).lower()


def _format_context(context_fields: dict[str, str], *, limit: int = 8) -> str:
    parts = [
        f"{key}={value}"
        for key, value in list(context_fields.items())[:limit]
        if key and value
    ]
    return "; ".join(parts)


def _answer_digest(answer: str, limit: int = 240) -> str:
    normalized = re.sub(r"\s+", " ", answer).strip()
    if len(normalized) <= limit:
        return normalized
    return normalized[:limit].rstrip() + "..."


def _slug(value: str) -> str:
    slug = re.sub(r"[^0-9A-Za-z\u4e00-\u9fff]+", "-", value.strip()).strip("-")
    return slug or "sheet"
